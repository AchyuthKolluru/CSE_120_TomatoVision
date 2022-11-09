import random
import openpyxl
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime 

# Global Variables
dt_string = datetime.now().strftime("%m/%d/%Y %H:%M:%S")

# Load in workbook
workbook = load_workbook(filename = 'testFile.xlsx')
ws = workbook.worksheets[0]

# Tomatoes (to be updated later after connected to other parts)
wholeTomatoes = 20 + random.randint(50, 100)
# slicedTomatoes = 20 + random.randint(50, 100)
# crushedTomatoes = 20 + random.randint(50, 100)

totalOfTomatoes = wholeTomatoes #+ slicedTomatoes + crushedTomatoes

def shift():
    # Updates shift column 
    for row in range(1 , 25):
        if(ws.cell(row,1).value is None):
            ws.cell(row,1).value = dt_string
            break

def whole_tomatoes():
    # For the whole tomatoes
    for row in range(1 , 25):
        if(ws.cell(row, 2).value is None):
            ws.cell(row, 2).value = wholeTomatoes
            break

def total_tomatoes():
    # For the total tomatoes
    for row in range(1 , 25):
        if(ws.cell(row,3).value is None):
            ws.cell(row,3).value = totalOfTomatoes
            break

def main():
    # Makes sure cell 'A1' is always 'Shift' in case someone accidentally writes into cell
    if ws.cell(1, 1).value == 'Shift':
        shift()
    else:
         ws.cell(1,1).value = 'Shift'
         shift()

    # Makes sure cell 'B1' is always 'Whole Tomatoes' 
    if ws.cell(1, 2).value == 'Whole Tomatoes':
        whole_tomatoes()
    else:
        ws.cell(1, 2).value = 'Whole Tomatoes'
        whole_tomatoes()

    # Makes sure cell 'C1' is always 'Total Tomatoes' 
    if ws.cell(1, 3).value == 'Total Tomatoes':
        total_tomatoes()
    else:
        ws.cell(1, 3).value = 'Total Tomatoes'
        total_tomatoes()

    workbook.save('testFile.xlsx')

if __name__ == "__main__":
    main()