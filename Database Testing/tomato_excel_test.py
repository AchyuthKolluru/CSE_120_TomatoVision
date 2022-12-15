import random
import openpyxl
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from datetime import datetime 

# Global Variables
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# Tomatoes (to  be updated later after connected to other parts)
wholeTomatoes = 20 + random.randint(50, 100)
# slicedTomatoes = 20 + random.randint(50, 100)
# crushedTomatoes = 20 + random.randint(50, 100)

totalOfTomatoes = wholeTomatoes #+ slicedTomatoes + crushedTomatoes

def warning(row, col, value):
    # Changes cell color to red if tomato amount is over 100
    # (to be changed later when mentor says what the value should be)
    if value > 100:
        ws.cell(row, col).fill = redFill

def shift():
    # Updates shift column 
    for row in range(1 , 25):
        if(ws.cell(row, 1).value is None):
            ws.cell(row, 1).value = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
            break

def whole_tomatoes():
    # For the whole tomatoes
    for row in range(1 , 25):
        if(ws.cell(row, 2).value is None):
            ws.cell(row, 2).value = wholeTomatoes
            warning(row, 2, ws.cell(row,2).value)
            break

def total_tomatoes():
    # For the total tomatoes
    for row in range(1 , 25):
        if(ws.cell(row,3).value is None):
            ws.cell(row,3).value = totalOfTomatoes
            break

def main():
    # Makes sure cell 'A1' is always 'Shift' in case someone accidentally writes into cell
    if ws.cell(1, 1).value == 'Shift':
        shift()
    else:
         ws.cell(1,1).value = 'Shift'
         shift()

    # Makes sure cell 'B1' is always 'Whole Tomatoes' 
    if ws.cell(1, 2).value == 'Whole Tomatoes':
        whole_tomatoes()
    else:
        ws.cell(1, 2).value = 'Whole Tomatoes'
        whole_tomatoes()

    # Makes sure cell 'C1' is always 'Total Tomatoes' 
    if ws.cell(1, 3).value == 'Total Tomatoes':
        total_tomatoes()
    else:
        ws.cell(1, 3).value = 'Total Tomatoes'
        total_tomatoes()

    workbook.save(file)

if __name__ == "__main__":
    # Load in workbook
    file = input("Input file name ([name].xlsx): ")
    workbook = load_workbook(file)
    ws = workbook.worksheets[0]
    main()